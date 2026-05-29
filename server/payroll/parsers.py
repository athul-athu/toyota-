from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

# Flexible column aliases (lowercase)
COLUMN_MAP = {
    "employee_id": ["employee id", "employee_id", "emp id", "emp_id", "id"],
    "name": ["name", "employee name", "employee_name"],
    "email": ["email", "e-mail", "mail"],
    "designation": ["designation", "title", "role", "position"],
    "base_salary": ["base salary", "base_salary", "basic", "basic salary"],
    "hra": ["hra", "house rent"],
    "allowances": ["allowances", "allowance", "other allowances"],
    "deductions": ["deductions", "deduction"],
    "period": [
        "month/year",
        "month_year",
        "month-year",
        "period",
        "pay period",
        "salary period",
    ],
    "month": ["month", "salary month"],
    "year": ["year", "salary year"],
}


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _map_headers(headers: list[str]) -> dict[str, int]:
    normalized = [_normalize_header(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, aliases in COLUMN_MAP.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                mapping[field] = idx
                break
    return mapping


def _to_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"Invalid number: {value}")


def _to_int(value: Any) -> int:
    if value is None or value == "":
        raise ValueError("Month/year is required")
    return int(float(str(value).strip()))


def _parse_period(value: Any, row_num: int) -> tuple[int, int]:
    """Parse month/year from values like 5/2026, 05/2026, or Excel dates."""
    if value is None or value == "":
        raise ValueError(f"Row {row_num}: Month/Year is required (format: month/year)")

    if hasattr(value, "month") and hasattr(value, "year"):
        return int(value.month), int(value.year)

    text = str(value).strip()
    for sep in ("/", "-", "."):
        if sep in text:
            parts = text.split(sep, 1)
            if len(parts) == 2:
                return _to_int(parts[0]), _to_int(parts[1])

    raise ValueError(
        f"Row {row_num}: Invalid Month/Year '{value}' — use format month/year (e.g. 5/2026)"
    )


def _resolve_month_year(row: list[Any], col_map: dict[str, int], row_num: int) -> tuple[int, int]:
    if "period" in col_map:
        idx = col_map["period"]
        value = row[idx] if idx < len(row) else None
        return _parse_period(value, row_num)
    if "month" in col_map and "year" in col_map:
        idx_m, idx_y = col_map["month"], col_map["year"]
        return _to_int(row[idx_m] if idx_m < len(row) else None), _to_int(
            row[idx_y] if idx_y < len(row) else None
        )
    raise ValueError(
        f"Row {row_num}: Missing Month/Year column — add 'Month/Year' (e.g. 5/2026) "
        "or separate Month and Year columns"
    )


def _parse_row(row: list[Any], col_map: dict[str, int], row_num: int) -> dict:
    required = [
        "employee_id",
        "name",
        "email",
        "base_salary",
        "hra",
        "allowances",
        "deductions",
    ]
    missing = [f for f in required if f not in col_map]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    if "period" not in col_map and ("month" not in col_map or "year" not in col_map):
        raise ValueError(
            "Missing period column: add 'Month/Year' (format month/year) or Month + Year columns"
        )

    def cell(field: str) -> Any:
        idx = col_map[field]
        return row[idx] if idx < len(row) else None

    employee_id = str(cell("employee_id") or "").strip()
    if not employee_id:
        raise ValueError(f"Row {row_num}: Employee ID is required")

    base = _to_decimal(cell("base_salary"))
    hra = _to_decimal(cell("hra"))
    allowances = _to_decimal(cell("allowances"))
    deductions = _to_decimal(cell("deductions"))
    net = base + hra + allowances - deductions
    month, year = _resolve_month_year(row, col_map, row_num)

    return {
        "employee_id": employee_id,
        "name": str(cell("name") or "").strip(),
        "email": str(cell("email") or "").strip(),
        "designation": str(cell("designation") or "").strip()
        if "designation" in col_map
        else "",
        "base_salary": float(base),
        "hra": float(hra),
        "allowances": float(allowances),
        "deductions": float(deductions),
        "net_salary": float(net),
        "month": month,
        "year": year,
        "period": f"{month}/{year}",
        "row": row_num,
    }


def _rows_from_csv(content: bytes) -> list[list[Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return list(reader)


def _rows_from_excel(content: bytes) -> list[list[Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def parse_payroll_file(filename: str, content: bytes) -> dict:
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = _rows_from_csv(content)
    elif lower.endswith((".xlsx", ".xls")):
        rows = _rows_from_excel(content)
    else:
        raise ValueError("Only CSV and Excel (.xlsx) files are supported")

    if len(rows) < 2:
        raise ValueError("File must have a header row and at least one data row")

    headers = [str(c) if c is not None else "" for c in rows[0]]
    col_map = _map_headers(headers)

    parsed: list[dict] = []
    errors: list[str] = []

    for i, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        try:
            parsed.append(_parse_row(list(row), col_map, i))
        except (ValueError, TypeError) as exc:
            errors.append(str(exc))

    if not parsed and errors:
        raise ValueError("; ".join(errors[:5]))

    return {
        "rows": parsed,
        "errors": errors,
        "count": len(parsed),
    }
